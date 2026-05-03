import os
import subprocess
import sys
from datetime import date

from openpyxl import Workbook

from backend.core import excel_analysis
from backend.core.file_access import inspect_file_path


def test_backend_startup_import_does_not_load_removed_excel_dependencies():
    env = os.environ.copy()
    env["PYTHONPATH"] = os.getcwd()
    result = subprocess.run(
        [
            sys.executable,
            "-c",
            "import sys; import backend.main; print(any(name in sys.modules for name in ('pandas', 'numpy', 'xlrd')))",
        ],
        cwd=os.getcwd(),
        env=env,
        text=True,
        capture_output=True,
        check=True,
    )

    assert result.stdout.strip() == "False"


def test_excel_inspect_reads_xlsx_without_loading_workbook_package(tmp_path):
    path = tmp_path / "sample.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["과제명", "담당자"])
    worksheet.append(["A", "Kim"])
    workbook.save(path)

    result = inspect_file_path(str(path))

    assert result["columns"] == ["과제명", "담당자"]


def test_excel_inspect_ignores_malformed_custom_properties(tmp_path):
    import zipfile

    path = tmp_path / "bad-custom-props.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["과제명", "담당자"])
    worksheet.append(["A", "Kim"])
    workbook.save(path)

    malformed_custom_props = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
  xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <property fmtid="{D5CDD505-2E9C-101B-9397-08002B2CF9AE}" pid="2">
    <vt:lpwstr>nameless</vt:lpwstr>
  </property>
</Properties>
"""
    with zipfile.ZipFile(path, "a") as archive:
        archive.writestr("docProps/custom.xml", malformed_custom_props)

    result = inspect_file_path(str(path))

    assert result["columns"] == ["과제명", "담당자"]


def test_excel_inspect_preserves_formatted_date_text(tmp_path):
    path = tmp_path / "dates.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["날짜", "값"])
    worksheet.append([date(2026, 4, 28), "마감"])
    workbook.save(path)

    result = inspect_file_path(str(path))

    assert result["sample"][0][0] == "2026-04-28"
    assert result["sample"][0][0] != "46140"


def test_excel_inspect_uses_first_non_empty_visible_sheet(tmp_path):
    path = tmp_path / "non-empty-second-sheet.xlsx"
    workbook = Workbook()
    workbook.active.title = "빈시트"
    data_sheet = workbook.create_sheet("데이터")
    data_sheet.append(["항목", "값"])
    data_sheet.append(["중요", "세컨드시트"])
    workbook.save(path)

    result = inspect_file_path(str(path))

    assert result["columns"] == ["항목", "값"]
    assert result["sample"][0] == ["중요", "세컨드시트"]


def test_extract_excel_used_ranges_reads_all_visible_sheets_and_skips_hidden(tmp_path):
    path = tmp_path / "all-visible-sheets.xlsx"
    workbook = Workbook()
    workbook.active.title = "요약"
    workbook.active["A1"] = "요약키워드"
    detail = workbook.create_sheet("세부")
    detail["B2"] = "세부키워드"
    hidden = workbook.create_sheet("숨김")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "숨겨진키워드"
    workbook.save(path)

    ranges = excel_analysis.extract_excel_used_ranges(str(path))

    assert [item.sheet_name for item in ranges] == ["요약", "세부"]
    assert ranges[1].value_at(1, 1) == "세부키워드"
    assert ranges[1].non_empty_cell_count == 1


def test_excel_discovery_keeps_full_used_range_for_large_sheet(tmp_path):
    path = tmp_path / "large.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["과제명", "담당자"])
    rows = 180
    for idx in range(rows):
        worksheet.append([f"A{idx}", "Kim"])
    workbook.save(path)

    result = inspect_file_path(str(path))
    ranges = excel_analysis.extract_excel_used_ranges(str(path))

    assert ranges[0].row_count == rows + 1
    assert result["sample"][0] == ["A0", "Kim"]
